import win32com.client
import tkinter as tk
from tkinter import filedialog as fd
import openpyxl
import os
from openpyxl.styles import PatternFill
from dictionary import fields, structures, markers
import os.path
import shutil
import tkinter.messagebox as mb
import re

# название формы
form_path = os.path.abspath('WellInfo.xlsx')
ppt_path = os.path.abspath('Sign Off Template.pptx')


# класс для создания экземпляра эксель-файла
class Workbook:
    def __init__(self, wb):
        self.wb = wb

    def load_wb(self):
        self.wb = openpyxl.load_workbook(self.wb)
        self.wb.active = 0
        self.ws = self.wb.active


class TkinterWindow:

    def callback(self):
        self.name = fd.askopenfilename()


# Создаём окно с выбором файла
tkinter = TkinterWindow()
my_w = tk.Tk()
my_w.geometry("400x200")  # Size of the window
exit_button = tk.Button(my_w, text="Закрыть", command=my_w.destroy)
exit_button.grid(row=2, column=0, padx=10, pady=20)
b1 = tk.Button(my_w, text='Выбери файл', bg='lightgreen', font=10,
               command=tkinter.callback, width=40)
b1.grid(row=0, column=0, padx=10, pady=20)
my_w.mainloop()

# Сохраняем название выбранного файла в переменную name
name = tkinter.name

# Открывает выбранный эксель файл
wb = openpyxl.load_workbook(name.split('/')[-1])
sheet = wb["Short"]
report = Workbook(name)
report.load_wb()
base_dir = os.path.dirname(form_path)
form_name = " ".join(name.split('/')[-1].split()[:4])
new_form_path = os.path.join(base_dir, f'{form_name}.xlsx')
# Создаём копию файла WellInfo
shutil.copy(form_path, new_form_path)
well_info = Workbook(new_form_path)
well_info.load_wb()


# вставляем название плана
well_info.ws['A4'].value = report.ws['C12'].value
well_name = report.ws['C9'].value
if "Well" in str(well_name):
    well_name = str(well_name).split()[1]


# вставляем название платформы
if len(report.ws['C8'].value.split()) > 0:
    structure = report.ws['C8'].value.split()[0]
else:
    structure = report.ws['C8'].value
for i in structures.items():
    if structure == i[0]:
        well_info.ws['A5'].value = i[1]


# вставляем название месторождения
if report.ws['C7'].value not in fields.keys():
    mb.showwarning(
        f"Внесите месторождение {report.ws['C7'].value}в словарик")
    print(f"Внесите месторождение  {report.ws['C7'].value}в словарик")
for i in fields.items():
    if report.ws['C7'].value == i[0]:
        well_info.ws['A3'].value = f"Месторождение {i[1]} Скважина {well_name}"


# вставляем альтитуда
rotary_table = (report.ws['I9'].value.split()[0])
sea_level = report.ws['I10'].value.split()[0]
rotary_sea_level = float(rotary_table) + float(sea_level)


# вставляем альтитуды в форму
well_info.ws['A6'].value = f"Глубина воды = {sea_level}м"
well_info.ws['A7'].value = f"Стол ротора = {rotary_table}м"
well_info.ws['A8'].value = f"Стол ротора - дно = {rotary_sea_level}м"
well_info.wb.save(new_form_path)


# вставляем таблицу в форму, оформляем и переводим
well_info.wb.active = 1
ws = well_info.wb.active
column_b = report.ws['B']
j = 2
len_column = len(column_b)

for i in range(23, len_column):
    if column_b[i].value != None:

        ws[f'A{j}'].value = report.ws['A'][i].value
        ws[f'B{j}'].value = report.ws['B'][i].value
        ws[f'C{j}'].value = report.ws['C'][i].value
        ws[f'D{j}'].value = report.ws['D'][i].value
        ws[f'E{j}'].value = report.ws['E'][i].value
        ws[f'F{j}'].value = report.ws['F'][i].value
        ws[f'G{j}'].value = report.ws['G'][i].value
        ws[f'H{j}'].value = report.ws['H'][i].value
        ws[f'I{j}'].value = report.ws['I'][i].value
        ws[f'J{j}'].value = report.ws['J'][i].value
        ws[f'K{j}'].value = report.ws['K'][i].value
        ws[f'L{j}'].value = report.ws['L'][i].value
        ws[f'M{j}'].value = report.ws['M'][i].value
        ws[f'N{j}'].value = report.ws['N'][i].value
        ws[f'O{j}'].value = report.ws['O'][i].value
        ws[f'P{j}'].value = report.ws['P'][i].value
        ws[f'Q{j}'].value = report.ws['Q'][i].value
        ws[f'R{j}'].value = report.ws['R'][i].value

        i = i + 1
        j = j+1
    else:
        break
# Проверяем наличие маркеров в словаре
unknown_markers = []
report_markers = []
for i in range(2, len(ws['A'])+1):
    if ws[f'A{i}'].value != None and ws[f'A{i}'].value != '' and not 'T' in ws[f'A{i}'].value:
        report_markers.append(ws[f'A{i}'].value)
        i = i+1
unknown_markers = [
    el for el in report_markers if el not in list(markers.keys())]

if len(unknown_markers) > 0:
    mb.showwarning('Занеси в словарик :',
                   (', '.join(map(str, unknown_markers))))
    print('Занеси в словарик маркер:', (', '.join(map(str, unknown_markers))))
# Переводим маркеры
for j in markers.items():
    for i in range(2, len(ws['A'])+1):
        if ws[f'A{i}'].value == j[0]:
            ws[f'A{i}'].value = j[1]

# Проверяю есть ли маркер 100% направленного бурения
marks = [ws[f'A{i}'].value for i in range(
    2, len(ws['A'])+1) if ws[f'A{i}'].value != None]

if '100% направленного бурения' not in marks:
    for i in range(2, len(ws['A'])+1):
        if ws[f'A{i}'].value == "Зумпф":
            if float(ws[f'B{i+1}'].value) - float(ws[f'B{i}'].value) == 20:
                ws[f'A{i+1}'].value = "100% направленного бурения"
            if float(ws[f'B{i+2}'].value) - float(ws[f'B{i}'].value) == 20:
                ws[f'A{i+2}'].value = "100% направленного бурения"


# #Закрашиваем строки
for i in range(2, len(ws['A'])+1):

    if "Башмак" in str(ws[f'A{i}'].value):
        range1 = ws[f'A{i}:R{i}']
        for cell in range1:
            for x in cell:
                x.fill = PatternFill('solid', fgColor="9CBBF0")
    if "Забой" in str(ws[f'A{i}'].value):
        range2 = ws[f'A{i}:R{i}']
        for cell in range2:
            for x in cell:
                x.fill = PatternFill('solid', fgColor="AAECA8")

# заполняем ControlList
well_info.wb.active = 2
ws2 = well_info.wb.active

for i in fields.items():
    if report.ws['C7'].value == i[0]:
        ws2['D4'].value = f'Месторождение {i[1]}'

slot = str(report.ws['C8'].value).split()[-1]
ws2['D6'].value = slot
ws2['D8'].value = well_name
ws2['D12'].value = report.ws['C12'].value

# Выясняем является ли скважина боковым стволом
reg_expr = r'(\BH[2-4])'
match = re.findall(reg_expr, str(report.ws['C12'].value))
reg_expr2 = r'(\ST)'
match2 = re.findall(reg_expr2, str(report.ws['C12'].value))
if len(match) > 0 or len(match2) > 0:

    ws2['D10'].value = 'Боковой Ствол'
    ws2['K17'].value = 'Нет'
    ws2['K18'].value = 'Да'
    ws2['I37'].value = "Н/П"
    ws2['I38'].value = "Н/П"
    ws2['I39'].value = "Н/П"
    ws2['K37'].value = "Н/П"
    ws2['K38'].value = "Н/П"
    ws2['K39'].value = "Н/П"

else:
    ws2['D10'].value = 'Основной Ствол'
    ws2['K17'].value = 'Да'
    ws2['K18'].value = 'Нет'

# Заполняем интенсивность
for i in range(2, len(ws['B'])):
    while ws[f'B{i}'].value != None:
        i = i+1
    else:
        break
last_row = i


def check_dls():
    interval = [j for j in range(2, last_row) if ws[f'B{j}'].value != None]
    ws2['K22'].value = 'Н/П'
    ws2['K23'].value = 'Н/П'
    ws2['K24'].value = 'Н/П'

    for i in interval:
        if ws[f'A{i}'].value == "100% направленного бурения":
            dls = [ws[f'O{j}'].value for j in range(
                ws[f'A{i}'].row+1, interval[-1]) if ws[f'O{j}'].value != 'N/A' and float(ws[f'O{j}'].value) > 2.5]
            if len(dls) > 0:
                ws2['K25'].value = "Нет"
            else:
                ws2['K25'].value = "Да"


if len(match) > 0 or len(match2) > 0:
    check_dls()
else:
    interval = [j for j in range(
        2, last_row-1) if ws[f'A{j}'].value != None and 'Башмак' in str(ws[f'A{j}'].value)]

    conductor_dls = [ws[f'O{j}'].value for j in range(
        interval[0]+1, interval[1]+1) if ws[f'O{j}'].value != 'N/A' and round((ws[f'O{j}'].value), 2) > 2]
    conductor_incl = [ws[f'C{j}'].value for j in range(
        2, interval[1]+1) if ws[f'C{j}'].value != 'N/A' and round((ws[f'C{j}'].value), 2) > 25]
    tech_casing_dls = [ws[f'O{j}'].value for j in range(
        interval[1]+1, interval[2]+1) if ws[f'O{j}'].value != 'N/A' and round((ws[f'O{j}'].value), 2) > 2]
    tech_casing_incl = [ws[f'C{j}'].value for j in range(
        interval[1]+1, interval[2]+1) if ws[f'C{j}'].value != 'N/A' and round((ws[f'C{j}'].value), 2) > 48]
    prod_casing_dls = [ws[f'O{j}'].value for j in range(
        interval[2]+1, interval[3]+1) if ws[f'O{j}'].value != 'N/A' and round((ws[f'O{j}'].value), 2) > 3]
    liner_dls = [ws[f'O{j}'].value for j in range(
        interval[3]+1, i) if ws[f'O{j}'].value != 'N/A' and round((ws[f'O{j}'].value), 2) > 2.5]
    print(tech_casing_dls)
    print(tech_casing_incl)
    if len(conductor_dls) or len(conductor_incl) > 0:
        ws2['K22'].value = 'Нет'
    else:
        ws2['K22'].value = 'Да'
    if len(tech_casing_dls) or len(tech_casing_incl) > 0:
        ws2['K23'].value = 'Нет'
    else:
        ws2['K23'].value = 'Да'

    if len(prod_casing_dls) > 0:
        ws2['K24'].value = 'Нет'
    else:
        ws2['K24'].value = 'Да'

    if len(liner_dls) > 0:
        ws2['K25'].value = 'Нет'
    else:
        ws2['K25'].value = 'Да'


# вставляю глубины колонн
for i in range(2, len(ws['A'])+1):
    if '508мм' in str(ws[f'A{i}'].value):
        ws2['I37'].value = ws[f'B{i}'].value
        ws2['K37'].value = ws[f'F{i}'].value

    if "Башмак колонны 339.7мм" in str(ws[f'A{i}'].value):

        ws2['I38'].value = ws[f'B{i}'].value
        ws2['K38'].value = ws[f'F{i}'].value
    if "Башмак колонны 406.4мм" in str(ws[f'A{i}'].value):

        ws2['I38'].value = ws[f'B{i}'].value
        ws2['K38'].value = ws[f'F{i}'].value

    if "Башмак колонны 244.5мм" in str(ws[f'A{i}'].value):

        ws2['I39'].value = ws[f'B{i}'].value
        ws2['K39'].value = ws[f'F{i}'].value

    if "Башмак колонны 273мм" in str(ws[f'A{i}'].value):
        ws2['I39'].value = ws[f'B{i}'].value
        ws2['K39'].value = ws[f'F{i}'].value

    if "Башмак колонны 273.05мм" in str(ws[f'A{i}'].value):
        print("yes")
        ws2['I39'].value = ws[f'B{i}'].value
        ws2['K39'].value = ws[f'F{i}'].value

len_column = len(ws['B'])

for l in range(2, len_column):
    if ws[f'B{l}'].value != None:
        l = l+1
    else:
        break
ws2['I40'].value = ws[f'B{l-1}'].value
ws2['K40'].value = ws[f'F{l-1}'].value
rng = ws[f'A{l-1}:R{l-1}']
for cell in rng:
    for x in cell:
        x.fill = PatternFill('solid', fgColor="AAECA8")
report.wb.close()
well_info.wb.save(new_form_path)

# Открываем powerpoint шаблон
pptApp = win32com.client.Dispatch('PowerPoint.Application')
ppt = pptApp.Presentations.Open(ppt_path, ReadOnly=False)

# Открываем заполненный файл WellInfo
excel_object = win32com.client.Dispatch("Excel.Application")
excel_workbook = excel_object.Workbooks.Open(Filename=new_form_path)

# Копируем таблицу и вставляем в слайд презентации
excel_worksheet = excel_workbook.Worksheets("RusTable")
excel_range = excel_worksheet.Range(f"A1:R{l-1}")
excel_range.Copy()
ppt.slides[0].Shapes.PasteSpecial(DataType=10, Link=True)
excel_object = win32com.client.Dispatch("Excel.Application")
excel_workbook = excel_object.Workbooks.Open(Filename=new_form_path)
excel_worksheet2 = excel_workbook.Worksheets("WellInfo")
excel_range2 = excel_worksheet2.Range("A2:A9")
excel_range2.Copy()
ppt.slides[0].Shapes.PasteSpecial(DataType=10, Link=True)
ppt.slides[1].Shapes.PasteSpecial(DataType=10, Link=True)
ppt.slides[2].Shapes.PasteSpecial(DataType=10, Link=True)

# Корректируем положение таблицы
ppt.slides[0].Shapes[3].width = 535
ppt.slides[0].Shapes[3].left = 2
ppt.slides[0].Shapes[3].top = 270
ppt.slides[0].Shapes[4].top = 120
ppt.slides[2].Shapes[2].top = 80
excel_object = win32com.client.Dispatch("Excel.Application")
excel_workbook = excel_object.Workbooks.Open(Filename=new_form_path)
excel_worksheet3 = excel_workbook.Worksheets("ControlList")
# Копируем контрольный лист
excel_range3 = excel_worksheet3.Range("A2:K40")
excel_range3.Copy()
# Вставляем в слайд
ppt.slides[3].Shapes.PasteSpecial(DataType=9, Link=True)
ppt.slides[3].Shapes[0].width = 480
ppt.slides[3].Shapes[0].left = 20
ppt.slides[3].Shapes[0].top = 30
excel_object = win32com.client.Dispatch("Excel.Application")
excel_workbook = excel_object.Workbooks.Open(Filename=name)
excel_worksheet = excel_workbook.Worksheets("Short")
# Копируем изображения траектории из рапорта из вкладки Short
for shape in excel_worksheet.Shapes:
    if shape.Name.startswith("Picture"):
        shape.Copy()
        ppt.slides[1].Shapes.Paste()

# Корректируем положение элементов на слайдах
ppt.slides[1].Shapes[1].width = 535
ppt.slides[1].Shapes[1].left = 2
ppt.slides[1].Shapes[1].top = 80
ppt.slides[1].Shapes[2].width = 400
ppt.slides[1].Shapes[2].top = 220
ppt.slides[1].Shapes[2].left = 50
ppt.slides[1].Shapes[3].width = 400
ppt.slides[1].Shapes[3].top = 460
ppt.slides[1].Shapes[3].left = 50

# Сохраняем презентацию в корневой папке
base_dir = os.path.dirname(ppt_path)
sign_off = f"Plan Sign off Rus_{ws2['D12'].value}.pptx"

ppt.SaveAs(os.path.join(base_dir, sign_off))
pptApp.Quit()
